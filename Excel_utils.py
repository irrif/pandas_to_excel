from typing import Union, List
import pandas as pd
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def excel_utils_helpme():
    print('Pour en savoir plus sur une fonction tapez : helpme(function)')
    print('Fonctions de la librairies Excel_utils :')
    print('get_index')
    print('get_coord')
    print('apply_font_to_cells')
    print('apply_font_to_multiple_sheets')
    print('apply_font')
    print('save_as_date')
    print('clear_existing_style')
    print('save_df_on_excel')

    
def helpme(function):
    print(f"{function.__name__} :\n{function.__doc__}")

    
def get_index(letter: str = "A", add_one: bool = False) -> int:
    """
    Retourne l'index associé à la colonne Excel
    
    Parameters
    ----------
    letter : str, default="A"
        Chaine de caractère comprenant une ou plusieurs lettres.
        
    add_one : bool, default=False
        Booléen qui ajoute 1 si l'on veut l'index compté à partir de 0 (False) ou de 1 (True).
    
    Returns
    -------
    int
    
    Example
    -------
    >>> get_index('E', False) -> 4
    >>> get_index('E', True) -> 5
    """

    index = 0
    for char in letter:
        index = index * 26 + (ord(char) - ord('A') + 1)
    
    if index > 16_384:
        raise ValueError("Column provided is out of Excel columns range. Excel columns ranges from A (1) to XFD (16_384)")
        
    else:
        return index + add_one
    

def index_from_letter(df: pd.DataFrame = None, col_name: str = "A") -> str:
    """
    Retourne la lettre de colonne Excel associé à l'index dans le dataframe.
    
    Parameters
    ----------
    df : pd.DataFrame
        Dataframe contenant la/les colonnes.
        
    col_name : str
        Nom de la colonne dans le DataFrame.
    
    Returns
    -------
    str
    
    Example
    -------
    >>> index_from_letter(df, 'Colonne_7') -> G
    """
    
    # Création d'un dictionnaire des lettres de l'alphabet en clé et de la position respective (commençant par 1)
    # 1:A, 2:B, ...  
    alphabet_dict = {}
    for idx, char in enumerate(range(65, 91), start=1):
        alphabet_dict[idx] = chr(char)
    
    if isinstance(col_name, str):
        col_index = df.columns.get_loc(col_name)
        
        # Check that column index is not greater than Excel max column index
        if col_index + 1 > 16_384:
            raise ValueError("Specified dataframe column will fall outside the bounds of an Excel file (column index > 16 384)")
    
        letter_to_add = ''
        if col_index > 25:
            letter_to_add = alphabet_dict[col_index // 25]
            col_index = col_index % 26

        return f"{letter_to_add}{chr(65 + col_index)}"
    

def get_column_character(df: pd.DataFrame, col_name: Union[str, list]) -> Union[str, list]:
    """
    Retourne la lettre de colonne Excel associé à l'index dans le dataframe
    
    Parameters :
    ------------
    - df : DataFrame
        Dataframe contenant la/les colonnes
    - col_name : str or list
        Nom ou liste des colonnes
    
    Returns :
    ---------
    str or list
    
    Example :
    ---------
    >>> get_column_character(df, 'Colonne_1') -> 'A'
    >>> get_column_character(df, ['Colonne_1', 'Colonne_27']) -> ['A', 'AB']
    """
    # Only one column provided
    if isinstance(col_name, str):
        return index_from_letter(df, col_name)
    
    # Multiple column provided
    elif isinstance(col_name, list):
        return [index_from_letter(df, col) for col in col_name]
        
    else:
        print('Erreur : col_name doit être une str ou une liste')
    
    
def get_coord(letter: str = "A", row: int = 1, add_one: bool = False) -> tuple:
    """
    Donne les coordonnées d'un point d'une grille Excel
    
    Parameters
    ----------
    letter : str, default="A"
        La lettre dont on veux l'index de la colonne
    row : int, default=1
        La ligne dont on veux l'index de ligne (en réalité row-1)
    add_one : bool, default=False
        Booléen qui ajoute 1 si l'on veut l'index compté à partir de 0 (False) ou de 1 (True)
    
    Returns
    -------
    tuple
    
    Example
    -------
    >>> get_coord('H', 4) -> (7, 3)
    >>> get_coord('H', 4, True) -> (8, 3)
    """
    
    return get_index(letter, add_one), row-1


#-----------------------------------------------------------------------------------------------------------------------------------#
#----------------------------------------- Modification des caractéristiques de l'écriture -----------------------------------------#
#-----------------------------------------------------------------------------------------------------------------------------------#

def apply_font_to_cells(writer: pd.ExcelWriter, sheet_name: str = 'Feuil1',
                        min_row: int = 1, max_row: int = 1, min_col: int = 1, max_col: int = 1,
                        font_name: str = 'Arial', font_size: int = 11, bold: bool = False, color: str = '000000'):
    """
    Applique une police (type, taille, gras et couleur) à une/des cellules d'une feuille uniquement.
    Pour plus d'informations voir la fonction "apply_font".
    Par défaut applique une police normal de taille 1 en noir.
    """
    
    # Set the working Excel file
    workbook = writer.book
    # Set the working sheet
    worksheet = writer.sheets[sheet_name]
    
    # Define custom font style
    custom_font = Font(name=font_name, size=font_size, bold=bold, color=color)
    
    # Apply custom font, by rows, to all cells encountered
    for row in worksheet.iter_rows(min_row, max_row, min_col, max_col):
        for cell in row:
            cell.font = custom_font
            
            
def apply_font_to_multiple_sheets(writer: pd.ExcelWriter, sheet_name: Union[str, list] = 'Feuil1',
                                  min_row: int = 1, max_row: Union[int, str] = 1, min_col: int = 1, max_col: Union[int, str] = 1,
                                  font_name: str = 'Arial', font_size: int = 11, bold: bool = False, color: str = '000000'):
    """
    Fonction qui boucle sur des feuilles d'un fichier excel afin d'appliquer un style de police précis.
    Pour plus d'informations voir la fonction "apply_font".
    Par défaut applique une police normal de taille 1 en noir.
    """
    
    # Apply to only one specified sheet
    if isinstance(sheet_name, str) and sheet_name.lower() != 'all':
        
        # Retrieves last row of the current worksheet
        if max_row.lower() == 'last':
            max_row = writer.sheets[sheet_name].max_row
        
        # Retrieves last col of the current worksheet
        if max_col.lower() == 'last':
            max_col = writer.sheets[sheet_name].max_column
        
        # Apply custom font
        apply_font_to_cells(writer, sheet_name,
                            min_row, max_row, min_col, max_col,
                            font_name, font_size, bold, color)
    
    # Apply to all sheets listed
    elif isinstance(sheet_name, list):
        
        # Applies up to last row and last column
        if max_row.lower() == 'last' and max_col.lower() == 'last':
            for sheets in sheet_name:
                max_row = writer.sheets[sheets].max_row
                max_col = writer.sheets[sheets].max_column
                
                apply_font_to_cells(writer, sheets,
                                min_row, max_row, min_col, max_col,
                                font_name, font_size, bold, color)
            return
        
        # Applies up to last row and specified column
        elif max_row.lower() == 'last':
            for sheets in sheet_name:
                max_row = writer.sheets[sheets].max_row
                
                apply_font_to_cells(writer, sheets,
                                min_row, max_row, min_col, max_col,
                                font_name, font_size, bold, color)
            return
        
        # Applied up to specified row and last column
        elif max_col.lower() == 'last':
            for sheets in sheet_name:
                max_col = writer.sheets[sheets].max_column
                
                apply_font_to_cells(writer, sheets,
                                min_row, max_row, min_col, max_col,
                                font_name, font_size, bold, color)
            return
                
        # Applies up to specified row and specified col
        else:
            apply_font_to_cells(writer, sheets,
                                min_row, max_row, min_col, max_col,
                                font_name, font_size, bold, color)
    
    # Applies to all workbook sheets
    elif sheet_name.lower() == 'all':
        
        # Applies up to last row and last column
        if max_row.lower() == 'last' and max_col.lower() == 'last':
            for sheets in writer.sheets:
                max_row = writer.sheets[sheets].max_row
                max_col = writer.sheets[sheets].max_column
                
                apply_font_to_cells(writer, sheets,
                                min_row, max_row, min_col, max_col,
                                font_name, font_size, bold, color)
            return
        
        # Applies up to last row and specified column
        elif max_row.lower() == 'last':
            for sheets in writer.sheets:
                max_row = writer.sheets[sheets].max_row
                
                apply_font_to_cells(writer, sheets,
                                min_row, max_row, min_col, max_col,
                                font_name, font_size, bold, color)
            return
        
        # Applied up to specified row and last column
        elif max_col.lower() == 'last':
            for sheets in writer.sheets:
                max_col = writer.sheets[sheets].max_column
                
                apply_font_to_cells(writer, sheets,
                                min_row, max_row, min_col, max_col,
                                font_name, font_size, bold, color)
            return
                
        # Applies up to specified row and specified col
        else:
            apply_font_to_cells(writer, sheets,
                                min_row, max_row, min_col, max_col,
                                font_name, font_size, bold, color)
            
    else:
        print('Error in sheet_name')


def apply_font(file: str = None, writer: pd.ExcelWriter = None, mode: str = 'a', sheets: Union[str, list] = 'Feuil1',
               min_row: int = 1, max_row: Union[int, str] = 1, min_col: int = 1, max_col: Union[int, str] = 1,
               font_name: str = 'Arial', font_size: int = 11, bold: bool = False, color: str = '000000'):
    """
    Applique une police (type, taille, gras et couleur) à une/des cellules, d'une ou plusieurs lignes, et d'une ou plusieurs colonnes.
    
    Il est possible soit :
    - De spécifier le fichier Excel que l'on souhaite modifier (file)
    - De spécifier l'objet ExcelWriter
    
    Parameters
    ----------
    file : str
        Chemin du fichier Excel
        
    writer : ExcelWriter
        ExcelWriter déjà configuré.
    
    mode : {'w', 'a'}, default='a'
        Mode d'écriture sur le fichier.
        
        - 'w' : Écriture, supprime tout ce qui était précédemment présent.
        - 'a' : Append, ajoute à la suite de ce qui est déjà présent.
    
    sheets : str or list, default='Feuil1'
        - 'Feuil1' : Applique uniquement à la feuille spécifiée
        - ['Feuil1', 'Feuil2'] : Applique aux feuilles spécifiées
        - 'All' : Applique à toutes les feuilles
        
    min_row : int, default=1
        Première ligne où appliquer
        
    max_row : int or str, default=1
        - 3 : Applique jusqu'à la ligne 3
        - 'last' : Applique jusqu'à la dernière ligne de la feuille
        
    min_col : int, default=1
        Première colonne où appliquer
        
    max_col : int or str, default=1
        - 3 : Applique jusqu'à la colonne 3
        - 'last' : Applique jusqu'à la dernière colonne de la feuille
        
    font_name : str, default='Arial'
        Nom de la police
        
    font_size : int, default=11
        Taille de la police
        
    bold : bool, default=False
        - True : En gras
        - False : En normal
        
    color : str, default='000000'
        Code hexadecimal de couleur
        - 000000 : Noir
        - FFFFFF : Blanc
    
    Returns
    -------
    None
    
    Example
    -------
    >>> apply_font()
    """      
    
    if file:
        try:
            with pd.ExcelWriter(path=file, mode=mode, engine='openpyxl', if_sheet_exists='overlay') as writer:
                apply_font_to_multiple_sheets(writer, sheets,
                                              min_row, max_row, min_col, max_col,
                                              font_name, font_size, bold, color)

        except FileNotFoundError:
            print("Le fichier spécifiée n'existe pas ou est introuvable")
    
    elif writer:
        apply_font_to_multiple_sheets(writer, sheets,
                                      min_row, max_row, min_col, max_col,
                                      font_name, font_size, bold, color)
            

#-----------------------------------------------------------------------------------------------------------------------------------#
#------------------------------------------ Application de styles de cellule particuliers ------------------------------------------#
#-----------------------------------------------------------------------------------------------------------------------------------#
def apply_date_style(writer: pd.ExcelWriter, sheet_name: str = 'Feuil1', date_cols: str = None,
                     min_row: int = 2, max_row: int = 1000):
    """
    Fonction qui permet d'appliquer le format date Excel (DD/MM/YYYY) à une colonne.
    
    Parameters
    ----------
    writer : ExcelWriter
        Objet qui permet d'écrire sur un fichier excel
        
    sheet_name : str, default='Feuil1'
        Nom de la feuille concernée
        
    date_cols : str
         Colonne sur laquelle appliquer le format date.
        
    min_row : default=2
        Ligne à partir de laquelle on applique la mise en forme.
        Le numéro de ligne correspond à celui d'Excel (il ne s'agit pas d'un index -> 1 = ligne 1)
        
    max_row : default=1000
        Ligne jusqu'à laquelle on applique la mise en forme
        Le numéro de ligne correspond à celui d'Excel (il ne s'agit pas d'un index -> 100 = ligne 100)
        
    Returns
    -------
    None
    
    Example
    -------
    >>> apply_date_style(writer, "Feuille1", "D", min_row=2, max_row=25)
    """
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
                    
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    col_index = get_index(date_cols, add_one=True)

    # applique le format Date aux cellules excel d'une colonne (avec openpyxl impossible de faire la colonne entière sans selectionner toutes les cellules non vides)
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=col_index, max_col=col_index):
        for cell in row:
            try:
                cell.style = 'date_style'
            except ValueError:
                cell.style = date_style


def save_as_date(writer: pd.ExcelWriter, sheet_name: str = 'Feuil1', date_cols: Union[str, tuple, list] = None, min_row: int=2, max_row: int=1000):
    """
    Fonction qui permet d'appliquer le format date Excel (DD/MM/YYYY) à une colonne.
    
    Parameters
    ----------
    writer : pd.ExcelWriter
        Objet qui permet d'écrire sur un fichier excel
        
    sheet_name : str, default='Feuil1'
        Nom de la feuille concernée
        
    date_cols : str, tuple or list
         Colonne ou collection de nom de colonnes (A, B, C, ...)
        
    min_row : int, default=2
        Ligne à partir de laquelle on applique la mise en forme.
        Le numéro de ligne correspond à celui d'Excel (il ne s'agit pas d'un index -> 1 = ligne 1)
        
    max_row : int, default=1000
        Ligne jusqu'à laquelle on applique la mise en forme
        Le numéro de ligne correspond à celui d'Excel (il ne s'agit pas d'un index -> 100 = ligne 100)
        
    Returns
    -------
    None
    
    Example
    -------
    >>> save_as_date(writer, 'Feuil1', "D", min_row=2, max_row=25)
    >>> save_as_date(writer, 'Feuil1', ('B', 'E'), min_row=2, max_row=50)
    >>> save_as_date(writer, 'Feuil1', ['A', 'C'], min_row=2, max_row=42)
    """
    
    if isinstance(col, str):
        apply_date_style(writer, sheet_name, date_cols, min_row, max_row)

    elif isinstance(date_cols, (tuple, list)):
        for col in date_cols:
            apply_date_style(writer, sheet_name, col, min_row, max_row)
        

def clear_existing_style(writer: pd.ExcelWriter, sheet_name: str = 'Feuil1',
                         min_row: int = 1, max_row: int = 1000, min_col: int = 1, max_col: int = 1000):
    """
    Fonction qui supprime un potentiel style de cellule préalablement existant sur des cellules spécifiées.
    
    Parameters
    ----------
    writer : pd.ExcelWrtier
        Objet qui permet d'écrire dans un fichier Excel.
        
    sheet_name : str, default='Feuil1'
        nom de la feuille que l'on veut modifier
        
    min_row : int, default=1
        ligne de début de modification
        
    max_row : int, default=1000
        ligne de fin de modification
        
    min_col : int, default=1
        colonne de début de modfication
        
    min_col : int, default=1000
        colonne de fin de modfication
        
    Returns
    -------
    None
    
    Example
    -------
    clear_existing_style(writer, "Feuille1", 1, 100, 1, 10)
    """
    
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.style = 'Normal'
            
            
def apply_style_to_headers(writer: pd.ExcelWriter, sheet_name: str = 'Feuil1',
                           list_of_headers: List[list[list, dict]] = None, df: pd.DataFrame = None):
    """
    Fonction qui permet d'appliquer un style personnalisé a la premiere cellule de chaque colonne, pour personnaliser l'en-tête.
    
    Parameters
    ----------
    writer : pd.ExcelWriter
        Objet qui permet d'écrire dans un fichier Excel.
        
    sheet_name : str, default='Feuil1'
        nom de la feuille que l'on veut modifier
        
    list_of_headers: list or tuple
        Liste (ou tuple) de liste du style : [[colonnes, header_style]]
        colonnes : Liste des colonnes contenant les headers à modifier.
        header_style : Dictionnaire contenant les paramètres du style de headers à appliquer.
        
    df : pd.DataFrame
        DataFrame contenant les colonnes, dont le nom sera le headers avec le style spécifié.
        
    Returns
    -------
    None
    
    Example
    -------
    >>> apply_style_to_headers(writer, 'Feuil1', [liste_colonne, header_params], df)
    
    **Multiple headers styles**
    
    >>> df = pd.DataFrame({'Type': ['Voiture', 'Voiture'],
    ...                    'Marque': ['Peugeot', 'Audi'],
    ...                    'Modèle': ['208', 'A3'],
    ...                    'Motorisation': ['Essence', 'Essence']})
    >>> df
       Type      Marque  Modèle  Motorisation
    0  Voiture  Peugeot     208       Essence
    1  Voiture     Audi      A3       Essence
    >>> first_style_columns = ['Type', 'Marque']
    >>> second_style_columns = ['Modèle', 'Motorisation']
    >>> header_params = {'name': 'existing', 'font_name':'Arial', 'font_size':9,
    ...                  'bold': True, 'font_color': 'FFFFFF', 'h_align':'center',
    ...                  'v_align':'center', 'wrap':True, 'start_color': '0b64a0',
    ...                  'end_color': '0b64a0', 'fill_type':'solid', 'column_height': 34.7}
    >>> apply_style_to_headers(writer, 'Feuil1',
    ...                        [[first_style_columns, header_params],
    ...                         [second_style_columns, header_params]],
    ...                        df)
    """

    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Iterates on the list containing: the list of columns + the dictionary of header parameters
    for list_cols_params in list_of_headers:
        # Columns list
        list_cols = list_cols_params[0]
        # Parameters of the headers associated with this list of columns, style dictionary parameter:value
        header_params = list_cols_params[1]
        # Header style names
        header_style_name = header_params['name']
        
        # For some reasons, when the named style already exists, it must be deleted before
        if header_style_name in workbook.named_styles:
            del workbook._named_styles[workbook.named_styles.index(header_style_name)]

        # Creates the header style for the list of columns
        header_style = NamedStyle(
            name=header_params['name'],
            font=Font(name=header_params['font_name'], size=header_params['font_size'], bold=header_params['bold'], color=header_params['font_color']),
            fill=PatternFill(start_color=header_params['start_color'], end_color=header_params['end_color'], fill_type=header_params['fill_type']),
            alignment=Alignment(horizontal=header_params['h_align'], vertical=header_params['v_align'], wrap_text=header_params['wrap'])
        )

        # Applies the header to specified cells
        for col in list_cols:
            # Finds the column index and adds 1 because Excel starts counting at 1 and not 0
            col_index = df.columns.get_loc(col) + 1 
            # First cell of the column
            header_cell = worksheet.cell(row=1, column=col_index) 

            # Defines the height of the first cell of the column
            worksheet.row_dimensions[1].height = header_params['column_height']

            try:
                # Applied header style
                header_cell.style = header_style.name    
                worksheet.column_dimensions[get_column_letter(col_index)].bestFit = True
                
            except ValueError:
                header_cell.style = header_style
                worksheet.column_dimensions[get_column_letter(col_index)].bestFit = True


def save_df_on_excel(df: pd.DataFrame, file: str, sheet_name = 'Feuil1', na_rep = 'NaN', columns: Union[str, list] = None, header: bool = True, index: bool = True, 
                     point: tuple = ('A', 1), mode = 'a', engine = 'openpyxl', ise = 'overlay', float_format = '%.2f', date_format: bool = False,
                     date_cols: list = None, header_format: bool = False, headers_list: list = None):
    """
    Sauvegarde un dataframe, ou une colonne du df, dans la colonne et à partir de la ligne spécifiée, du fichier spécifié.
    
    Parameters
    ----------
    df : pd.DataFrame
        Le DataFrame à sauvegarder
        
    file : str or path
        Le fichier Excel dans lequel on veut sauvegarder
        
    sheet_name : str, default="Feuil1"
        Le nom de la feuille au sein de ce dit fichier
        
    na_rep : str, default='NaN'
        Façon dont sont représentées les valeurs manquantes dans le fichier Excel
        
        - '' : Si l'on souhaite laisser les cellules vides
        
    columns : str or list
        Colonne du DataFrame à écrire dans le fichier Excel
        
    header : bool, default=True
        Booleen qui renseigne ou non le nom des colonnes dans le fichier Excel
    
    index : bool, default=True
        Booleen qui renseigne ou non l'index dans le fichier Excel
        
    point : tuple, default=('A', 1)
        Coordonnées à partir duquel on sauvegarde le DataFrame
        
    mode : {'w', 'a'}, default='a'
        Mode d'écriture dans le fichier 
         
        - 'w' : écriture 
        - 'a' : append
        
    engine : str, default='openpyxl'
        openpyxl ou XlsxWriter
        
    ise : {'error', 'new', 'replace', 'overlay'}, default='overlay'
        Action à faire si mode=appending et que la feuille existe déjà
        
        - error: raise a ValueError.
        - new: Create a new sheet, with a name determined by the engine.
        - replace: Delete the contents of the sheet before writing to it.
        - overlay: Write contents to the existing sheet without removing
        
    date_format : bool, default=False
        Si l'on souhaite le format Date sur certaines colonnes du fichier Excel
        
    date_cols : list
        Liste des colonnes à transformer en format Date.
        Cette liste doit contenir la lettre de la colonne Excel (A, B, ...)
        
    header_format : bool, default=False
        Si l'on souhaite appliquer un style personnalisé ou non aux en-têtes de lignes
        
    headers_list : list
        Il s'agit d'une liste, contenant une liste, qui contient elle même une liste de colonne et un dictionnaire de mise en forme style CSS
    
    Returns
    -------
    None
    
    Example
    -------
    >>> 
    """
    
    col, row = get_coord(point[0], point[1])
    
    if not file.endswith(('.xlsx', '.xlsm')):
        print(f"The filename {file} does not provide extension, by default the extension '.xlsx' is used")
        file += '.xlsx'
    
    if mode == 'w':
        with pd.ExcelWriter(path=file, mode=mode, engine=engine) as writer:
            # Saves the dataframe on the Excel file
            df.to_excel(excel_writer=writer, sheet_name=sheet_name, na_rep=na_rep, columns=columns, header=header, index=index, startcol=col, startrow=row, float_format=float_format)
            
            # Transforms columns to date format if specified
            if date_format:
                save_as_date(writer=writer, sheet_name=sheet_name, date_cols=date_cols, min_row=2, max_row=len(df)+1)
            
            # Formats headers
            if header_format:
                # First, previous style has to be deleted
                clear_existing_style(writer, sheet_name, min_row=1, max_row=1, min_col=1, max_col=df.shape[1]+1)
                # Then applies current header style
                apply_style_to_headers(writer, sheet_name, headers_list, df)
            
            print('Written')

    elif mode == 'a':
        try:
            with pd.ExcelWriter(path=file, mode=mode, engine=engine, if_sheet_exists=ise) as writer:
                df.to_excel(excel_writer=writer, sheet_name=sheet_name, na_rep=na_rep, columns=columns, header=header, index=index, startcol=col, startrow=row, float_format=float_format)

                # Transforms columns to date format if specified
                if date_format:
                    save_as_date(writer=writer, sheet_name=sheet_name, date_cols=date_cols, min_row=2, max_row=len(df)+1)
                
                # Formats headers
                if header_format:
                    # First, previous style has to be deleted
                    clear_existing_style(writer, sheet_name, min_row=1, max_row=1, min_col=1, max_col=df.shape[1]+1)
                    # Then applies current header style
                    apply_style_to_headers(writer, sheet_name, headers_list, df)
                            
                print('Appended')
                
        # If the file doesn't exists, then runs the function in writting mode
        except FileNotFoundError:
            save_df_on_excel(df, file, sheet_name, na_rep, columns, header, index, point, mode='w', float_format=float_format,
                             date_format=date_format, date_cols=date_cols,
                             header_format=header_format, headers_list=headers_list)

    else:
        print(f"Erreur dans le mode spécifié : {mode} n'existe pas")